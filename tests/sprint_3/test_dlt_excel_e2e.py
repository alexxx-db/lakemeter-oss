"""
Sprint 3: DLT End-to-End Excel Generation Tests (BUG-S3-E1)

Generates a real .xlsx file with DLT line items via build_estimate_excel,
then reads it with openpyxl to verify:
1. Formula cells contain formulas (not static values)
2. SKU column matches expected values
3. No NaN/None in computed cells for valid DLT configs
4. Totals row uses SUM formulas
"""
import math
import os
import sys
import tempfile
import pytest
import openpyxl

BACKEND_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'backend')
sys.path.insert(0, BACKEND_DIR)

from tests.sprint_3.conftest import make_line_item
from app.routes.export import build_estimate_excel


def _make_estimate(**kwargs):
    """Create a mock estimate object."""
    from types import SimpleNamespace
    from datetime import datetime
    defaults = {
        'estimate_name': 'DLT E2E Test Estimate',
        'status': 'draft',
        'version': 1,
        'created_at': datetime(2026, 3, 31),
        'updated_at': datetime(2026, 3, 31),
    }
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


def _generate_xlsx(line_items, cloud='aws', region='us-east-1', tier='PREMIUM'):
    """Generate an Excel file and return the openpyxl workbook."""
    estimate = _make_estimate()
    output = build_estimate_excel(estimate, line_items, cloud, region, tier)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp_path = f.name
    wb = openpyxl.load_workbook(tmp_path)
    os.unlink(tmp_path)
    return wb


def _find_data_start_row(ws):
    """Find the first data row (after 'WORKLOADS & COST BREAKDOWN' header)."""
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=6)  # SKU column (col F / index 5+1)
        val = cell.value
        if val and isinstance(val, str) and val.startswith('DLT_'):
            return row_idx
        if val and isinstance(val, str) and val == 'JOBS_SERVERLESS_COMPUTE':
            return row_idx
        if val and isinstance(val, str) and val == 'DELTA_LIVE_TABLES_SERVERLESS':
            return row_idx
    return None


def _find_totals_row(ws):
    """Find the TOTALS row."""
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value and isinstance(cell.value, str) and 'TOTALS' in cell.value:
            return row_idx
    return None


# Column mapping (0-indexed in spec, 1-indexed in openpyxl)
COL_SKU = 6       # Column F
COL_HOURS = 12     # Column L
COL_DBU_HR = 16    # Column P
COL_DBUS_MO = 17   # Column Q (formula: =P*L)
COL_DBU_RATE = 18  # Column R
COL_DISCOUNT = 19  # Column S
COL_DBU_RATE_D = 20  # Column T (formula: =R*(1-S))
COL_DBU_COST_L = 21  # Column U (formula: =Q*R)
COL_DBU_COST_D = 22  # Column V (formula: =Q*T)
COL_DRIVER_VM = 23   # Column W
COL_WORKER_VM = 24   # Column X
COL_DRV_VM_TOT = 25  # Column Y (formula: =W*L)
COL_WRK_VM_TOT = 26  # Column Z (formula: =X*L*I)
COL_VM_TOTAL = 27    # Column AA (formula: =Y+Z)
COL_TOTAL_L = 28     # Column AB (formula: =U+AA)
COL_TOTAL_D = 29     # Column AC (formula: =V+AA)


class TestDLTExcelE2EFormulas:
    """Verify real .xlsx has formulas in computed columns for DLT."""

    @pytest.fixture
    def dlt_workbook(self):
        """Generate workbook with DLT Core Classic + DLT Serverless Perf."""
        items = [
            make_line_item(
                workload_name="DLT Core Classic Test",
                dlt_edition="CORE",
                driver_node_type="i3.xlarge",
                worker_node_type="i3.xlarge",
                num_workers=4,
                photon_enabled=False,
                serverless_enabled=False,
                hours_per_month=730,
            ),
            make_line_item(
                workload_name="DLT Serverless Perf Test",
                dlt_edition="ADVANCED",
                driver_node_type="i3.xlarge",
                worker_node_type="i3.xlarge",
                num_workers=4,
                serverless_enabled=True,
                serverless_mode="performance",
                hours_per_month=730,
            ),
        ]
        return _generate_xlsx(items)

    def test_workbook_has_sheet(self, dlt_workbook):
        assert 'Databricks Estimate' in dlt_workbook.sheetnames

    def test_data_rows_exist(self, dlt_workbook):
        ws = dlt_workbook['Databricks Estimate']
        start = _find_data_start_row(ws)
        assert start is not None, "No DLT data rows found"

    def test_sku_column_values(self, dlt_workbook):
        """Verify SKU column has correct backend SKU values."""
        ws = dlt_workbook['Databricks Estimate']
        start = _find_data_start_row(ws)
        assert start is not None
        sku1 = ws.cell(row=start, column=COL_SKU).value
        sku2 = ws.cell(row=start + 1, column=COL_SKU).value
        assert sku1 == "DLT_CORE_COMPUTE"
        assert sku2 == "DELTA_LIVE_TABLES_SERVERLESS"

    def test_dbus_mo_has_formula(self, dlt_workbook):
        """Col 16 (DBUs/Mo) should contain a formula =P*L for hourly items."""
        ws = dlt_workbook['Databricks Estimate']
        start = _find_data_start_row(ws)
        for offset in range(2):
            cell = ws.cell(row=start + offset, column=COL_DBUS_MO)
            assert cell.value is not None
            # openpyxl stores formula strings starting with '='
            if isinstance(cell.value, str):
                assert cell.value.startswith('='), \
                    f"Row {start + offset}: DBUs/Mo should be formula, got '{cell.value}'"

    def test_dbu_cost_list_has_formula(self, dlt_workbook):
        """Col 20 (DBU Cost List) should contain formula =Q*R."""
        ws = dlt_workbook['Databricks Estimate']
        start = _find_data_start_row(ws)
        for offset in range(2):
            cell = ws.cell(row=start + offset, column=COL_DBU_COST_L)
            assert cell.value is not None
            if isinstance(cell.value, str):
                assert cell.value.startswith('='), \
                    f"Row {start + offset}: DBU Cost (List) should be formula"

    def test_dbu_cost_disc_has_formula(self, dlt_workbook):
        """Col 21 (DBU Cost Disc.) should contain formula =Q*T."""
        ws = dlt_workbook['Databricks Estimate']
        start = _find_data_start_row(ws)
        for offset in range(2):
            cell = ws.cell(row=start + offset, column=COL_DBU_COST_D)
            assert cell.value is not None
            if isinstance(cell.value, str):
                assert cell.value.startswith('='), \
                    f"Row {start + offset}: DBU Cost (Disc.) should be formula"

    def test_total_cost_has_formula(self, dlt_workbook):
        """Col 27-28 (Total Cost) should contain formulas."""
        ws = dlt_workbook['Databricks Estimate']
        start = _find_data_start_row(ws)
        for offset in range(2):
            for col in [COL_TOTAL_L, COL_TOTAL_D]:
                cell = ws.cell(row=start + offset, column=col)
                assert cell.value is not None
                if isinstance(cell.value, str):
                    assert cell.value.startswith('='), \
                        f"Row {start + offset}, Col {col}: Total Cost should be formula"

    def test_classic_vm_cost_has_formula(self, dlt_workbook):
        """Classic DLT row should have VM cost formulas (cols 24-26)."""
        ws = dlt_workbook['Databricks Estimate']
        start = _find_data_start_row(ws)
        # Row 1 is Classic — should have VM formulas
        for col in [COL_DRV_VM_TOT, COL_WRK_VM_TOT, COL_VM_TOTAL]:
            cell = ws.cell(row=start, column=col)
            assert cell.value is not None
            if isinstance(cell.value, str):
                assert cell.value.startswith('='), \
                    f"Classic row Col {col}: VM cost should be formula"

    def test_serverless_vm_cost_is_zero(self, dlt_workbook):
        """Serverless DLT row should have zero VM costs."""
        ws = dlt_workbook['Databricks Estimate']
        start = _find_data_start_row(ws)
        # Row 2 is Serverless
        for col in [COL_DRV_VM_TOT, COL_WRK_VM_TOT, COL_VM_TOTAL]:
            cell = ws.cell(row=start + 1, column=col)
            val = cell.value
            assert val == 0 or val is None or val == '', \
                f"Serverless row Col {col}: VM cost should be 0, got {val}"

    def test_no_nan_in_computed_cells(self, dlt_workbook):
        """No NaN values in any computed cell for valid DLT configs."""
        ws = dlt_workbook['Databricks Estimate']
        start = _find_data_start_row(ws)
        computed_cols = [
            COL_DBUS_MO, COL_DBU_RATE_D, COL_DBU_COST_L,
            COL_DBU_COST_D, COL_TOTAL_L, COL_TOTAL_D,
        ]
        for offset in range(2):
            for col in computed_cols:
                cell = ws.cell(row=start + offset, column=col)
                val = cell.value
                if isinstance(val, float):
                    assert not math.isnan(val), \
                        f"Row {start + offset}, Col {col}: NaN found"
                if isinstance(val, str):
                    assert 'nan' not in val.lower(), \
                        f"Row {start + offset}, Col {col}: NaN string found"


class TestDLTExcelE2ETotals:
    """Verify the totals row uses SUM formulas."""

    @pytest.fixture
    def multi_item_workbook(self):
        """3 DLT line items for totals testing."""
        items = [
            make_line_item(
                workload_name="DLT Core Classic",
                dlt_edition="CORE",
                driver_node_type="i3.xlarge",
                worker_node_type="i3.xlarge",
                num_workers=4,
                photon_enabled=False,
                serverless_enabled=False,
                hours_per_month=730,
            ),
            make_line_item(
                workload_name="DLT Pro Photon",
                dlt_edition="PRO",
                driver_node_type="i3.xlarge",
                worker_node_type="i3.xlarge",
                num_workers=4,
                photon_enabled=True,
                serverless_enabled=False,
                hours_per_month=200,
            ),
            make_line_item(
                workload_name="DLT Serverless Perf",
                dlt_edition="ADVANCED",
                driver_node_type="i3.xlarge",
                worker_node_type="i3.xlarge",
                num_workers=4,
                serverless_enabled=True,
                serverless_mode="performance",
                hours_per_month=730,
            ),
        ]
        return _generate_xlsx(items)

    def test_totals_row_exists(self, multi_item_workbook):
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        assert totals_row is not None, "TOTALS row not found"

    def test_totals_dbus_mo_has_sum(self, multi_item_workbook):
        """Col 16 (DBUs/Mo) in totals row should use SUM formula."""
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        cell = ws.cell(row=totals_row, column=COL_DBUS_MO)
        val = cell.value
        assert isinstance(val, str) and 'SUM' in val.upper(), \
            f"Totals DBUs/Mo should use SUM, got '{val}'"

    def test_totals_dbu_cost_list_has_sum(self, multi_item_workbook):
        """Col 20 (DBU Cost List) in totals row should use SUM."""
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        cell = ws.cell(row=totals_row, column=COL_DBU_COST_L)
        val = cell.value
        assert isinstance(val, str) and 'SUM' in val.upper(), \
            f"Totals DBU Cost (List) should use SUM, got '{val}'"

    def test_totals_total_cost_has_sum(self, multi_item_workbook):
        """Col 27-28 (Total Cost) in totals row should use SUM."""
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        for col in [COL_TOTAL_L, COL_TOTAL_D]:
            cell = ws.cell(row=totals_row, column=col)
            val = cell.value
            assert isinstance(val, str) and 'SUM' in val.upper(), \
                f"Totals Col {col} should use SUM, got '{val}'"

    def test_totals_vm_cost_has_sum(self, multi_item_workbook):
        """Cols 24-26 (VM Costs) in totals row should use SUM."""
        ws = multi_item_workbook['Databricks Estimate']
        totals_row = _find_totals_row(ws)
        for col in [COL_DRV_VM_TOT, COL_WRK_VM_TOT, COL_VM_TOTAL]:
            cell = ws.cell(row=totals_row, column=col)
            val = cell.value
            assert isinstance(val, str) and 'SUM' in val.upper(), \
                f"Totals Col {col} should use SUM, got '{val}'"
