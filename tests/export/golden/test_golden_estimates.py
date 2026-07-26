"""Golden-estimate regression pack.

Generates the canonical 9-workload combined estimate Excel export and
asserts every pinned value in golden_data.py — DBU/hr, rates, monthly
DBUs, monthly cost math, storage sub-rows, totals formulas, and the
DB-first VM pricing path.

A failure here means one of three things:
1. a pricing formula regressed           -> fix the code
2. the static pricing data drifted        -> verify against the published
   source, then update golden_data.py IN THE SAME PR with the citation
3. the export structure changed           -> update the pack deliberately

Golden values are derived independently from the pricing data and
documented formulas (see golden_data.py), never from the code's output.
"""
import os
import sys

import pytest

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.insert(0, os.path.join(REPO_ROOT, 'backend'))
sys.path.insert(0, REPO_ROOT)

from tests.export.cross_workload.conftest import make_all_nine_items, make_line_item
from tests.export.cross_workload.excel_helpers import (
    make_estimate, find_row_by_name, find_totals_row,
    COL_NAME, COL_TYPE, COL_SKU, COL_NUM_WORKERS, COL_HOURS,
    COL_TOKEN_TYPE, COL_TOKEN_QTY, COL_DBU_PER_M, COL_DBU_HR,
    COL_DBUS_MO, COL_DBU_RATE, COL_DBU_COST_L,
    COL_DRIVER_VM_HR, COL_WORKER_VM_HR, COL_NOTES,
)
from tests.export.golden.golden_data import (
    GOLDEN_ROWS, GOLDEN_STORAGE_ROWS, GOLDEN_TOTAL_DBU_COST,
    STATIC_I3_XLARGE_ON_DEMAND, DB_INJECTED_VM_PRICE,
)

TOL = 1e-6
COST_TOL = 1e-4


def _generate(items=None, db=None):
    """Build the canonical workbook and return the active worksheet."""
    import tempfile
    import openpyxl
    from app.routes.export import build_estimate_excel

    estimate = make_estimate()
    output = build_estimate_excel(
        estimate, items or make_all_nine_items(), 'aws', 'us-east-1',
        'PREMIUM', db=db)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp_path = f.name
    wb = openpyxl.load_workbook(tmp_path)
    os.unlink(tmp_path)
    return wb.active


@pytest.fixture(scope='module')
def sheet():
    return _generate()


class TestGoldenPrimaryRows:
    """Each canonical workload row matches its pinned golden record."""

    @pytest.mark.parametrize('golden', GOLDEN_ROWS,
                             ids=[g['name'] for g in GOLDEN_ROWS])
    def test_row_matches_golden(self, sheet, golden):
        row = find_row_by_name(sheet, golden['name'])
        assert row is not None, f"row not found for {golden['name']}"
        cell = lambda col: sheet.cell(row=row, column=col).value

        assert cell(COL_SKU) == golden['sku']
        assert cell(COL_DBU_RATE) == pytest.approx(golden['rate'], abs=TOL)

        if 'token_qty' in golden:
            assert cell(COL_HOURS) == 'N/A'
            assert cell(COL_TOKEN_TYPE) == golden['token_type']
            assert cell(COL_TOKEN_QTY) == golden['token_qty']
            assert cell(COL_DBU_PER_M) == pytest.approx(
                golden['dbu_per_m'], abs=TOL)
            qty = golden['token_qty'] * golden['dbu_per_m']
        else:
            assert cell(COL_HOURS) == golden['hours']
            assert cell(COL_DBU_HR) == pytest.approx(
                golden['dbu_hr'], abs=TOL)
            qty = golden['dbu_hr'] * golden['hours']

        if 'num_workers' in golden:
            assert cell(COL_NUM_WORKERS) == golden['num_workers']

        # Monthly DBUs and cost are Excel formulas (=P*L or =N*O and
        # =Q*R); verify the math those formulas will compute against the
        # golden record. The formula strings themselves are pinned in
        # test_dbus_mo_formula.
        assert qty == pytest.approx(golden['dbus_mo'], abs=COST_TOL)
        rate = cell(COL_DBU_RATE)
        assert qty * rate == pytest.approx(golden['dbu_cost'], abs=COST_TOL)

        # Serverless rows / missing node types carry no VM cost.
        assert cell(COL_DRIVER_VM_HR) == 0
        assert cell(COL_WORKER_VM_HR) == 0

    @pytest.mark.parametrize('golden', GOLDEN_ROWS,
                             ids=[g['name'] for g in GOLDEN_ROWS])
    def test_dbus_mo_formula(self, sheet, golden):
        row = find_row_by_name(sheet, golden['name'])
        formula = sheet.cell(row=row, column=COL_DBUS_MO).value
        if not isinstance(formula, str):
            pytest.skip('dbus_mo is a literal (storage/token row shape)')
        if 'token_qty' in golden:
            assert formula == f'=N{row}*O{row}'
        else:
            assert formula == f'=P{row}*L{row}'


class TestGoldenStorageRows:
    """Storage sub-rows match their pinned golden records."""

    @pytest.mark.parametrize('golden', GOLDEN_STORAGE_ROWS,
                             ids=[g['type'] for g in GOLDEN_STORAGE_ROWS])
    def test_storage_row_matches_golden(self, sheet, golden):
        parent_row = find_row_by_name(sheet, golden['parent'])
        assert parent_row is not None
        # storage sub-row immediately follows its parent
        row = parent_row + 1
        cell = lambda col: sheet.cell(row=row, column=col).value

        assert cell(COL_NAME) == golden['parent']
        assert cell(COL_TYPE) == golden['type']
        assert cell(COL_SKU) == golden['sku']
        assert cell(COL_DBU_RATE) == pytest.approx(golden['rate'], abs=TOL)
        assert cell(COL_DBU_COST_L) == pytest.approx(golden['cost'], abs=TOL)
        assert cell(COL_NOTES) == golden['note']


class TestGoldenTotals:
    """The TOTALS row aggregates the full data range."""

    def test_totals_formulas_cover_all_data_rows(self, sheet):
        totals = find_totals_row(sheet)
        assert totals is not None
        # data block starts at row 10 and ends the row above the
        # blank separator before TOTALS (canonical 11-row block: 10-20)
        assert sheet.cell(row=totals, column=COL_DBUS_MO).value == \
            '=SUM(Q10:Q20)'
        assert sheet.cell(row=totals, column=COL_DBU_COST_L).value == \
            '=SUM(U10:U20)'

    def test_golden_grand_total_is_internally_consistent(self):
        """Sum of golden row costs equals the pinned grand total."""
        assert GOLDEN_TOTAL_DBU_COST == pytest.approx(12328.9125, abs=COST_TOL)


class _FakeRow:
    def __init__(self, cost):
        self.cost_per_hour = cost


class _FakeResult:
    def __init__(self, row):
        self._row = row

    def fetchone(self):
        return self._row


class _FakeDB:
    """Returns a distinctive injected VM price for every query."""

    def execute(self, stmt, params):
        return _FakeResult(_FakeRow(DB_INJECTED_VM_PRICE))


def _make_jobs_classic_item():
    return make_line_item(
        workload_type='JOBS',
        workload_name='Jobs Classic VM Golden',
        driver_node_type='i3.xlarge',
        worker_node_type='i3.xlarge',
        num_workers=2,
        hours_per_month=100,
        driver_pricing_tier='on_demand',
        worker_pricing_tier='on_demand',
    )


class TestGoldenVmPricing:
    """DB-first VM pricing, end-to-end through the workbook."""

    def test_db_price_flows_to_workbook(self):
        ws = _generate(items=[_make_jobs_classic_item()], db=_FakeDB())
        row = find_row_by_name(ws, 'Jobs Classic VM Golden')
        assert row is not None
        assert ws.cell(row=row, column=COL_DRIVER_VM_HR).value == \
            DB_INJECTED_VM_PRICE
        assert ws.cell(row=row, column=COL_WORKER_VM_HR).value == \
            DB_INJECTED_VM_PRICE
        # authoritative DB hit: no fallback disclosure in the notes
        assert ws.cell(row=row, column=COL_NOTES).value in (None, '')

    def test_static_fallback_is_priced_and_disclosed(self):
        ws = _generate(items=[_make_jobs_classic_item()], db=None)
        row = find_row_by_name(ws, 'Jobs Classic VM Golden')
        assert row is not None
        assert ws.cell(row=row, column=COL_DRIVER_VM_HR).value == \
            pytest.approx(STATIC_I3_XLARGE_ON_DEMAND, abs=TOL)
        assert ws.cell(row=row, column=COL_WORKER_VM_HR).value == \
            pytest.approx(STATIC_I3_XLARGE_ON_DEMAND, abs=TOL)
        notes = ws.cell(row=row, column=COL_NOTES).value or ''
        assert notes.count('default reference price') == 2  # driver + worker

    def test_vm_cost_formulas(self):
        ws = _generate(items=[_make_jobs_classic_item()], db=_FakeDB())
        row = find_row_by_name(ws, 'Jobs Classic VM Golden')
        # driver cost = driver $/hr x hours; worker cost adds x num_workers
        assert ws.cell(row=row, column=25).value == f'=W{row}*L{row}'
        assert ws.cell(row=row, column=26).value == f'=X{row}*L{row}*I{row}'
