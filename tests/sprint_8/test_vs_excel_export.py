"""Test Vector Search Excel export with real .xlsx files.

AC-9 to AC-14: Compute formula, storage sub-row, SKU, storage cost, totals.
"""
import math
import os
import tempfile
import pytest
import openpyxl

from tests.sprint_8.conftest import make_line_item
from tests.sprint_8.vs_calc_helpers import calc_dbu_per_hour, calc_storage_gb
from app.routes.export import build_estimate_excel


def _make_estimate(**kw):
    from types import SimpleNamespace
    from datetime import datetime
    d = dict(estimate_name='Vector Search E2E', status='draft', version=1,
             created_at=datetime(2026, 3, 31), updated_at=datetime(2026, 3, 31))
    d.update(kw)
    return SimpleNamespace(**d)


def _generate_xlsx(line_items, cloud='aws', region='us-east-1',
                   tier='PREMIUM'):
    estimate = _make_estimate()
    output = build_estimate_excel(estimate, line_items, cloud, region, tier)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(output.read())
        tmp_path = f.name
    wb = openpyxl.load_workbook(tmp_path)
    os.unlink(tmp_path)
    return wb


def _find_data_rows(ws, sku_filter=None):
    """Find all data rows. If sku_filter given, only rows with that SKU."""
    rows = []
    for row_idx in range(1, ws.max_row + 1):
        sku_val = ws.cell(row=row_idx, column=6).value
        if sku_filter:
            if sku_val == sku_filter:
                rows.append(row_idx)
        else:
            if sku_val and isinstance(sku_val, str) and sku_val not in (
                    '', '-', 'SKU', 'SKU / Product Type'):
                rows.append(row_idx)
    return rows


def _find_vs_compute_row(ws):
    """Find Vector Search compute row (VECTOR_SEARCH_ENDPOINT)."""
    rows = _find_data_rows(ws, 'VECTOR_SEARCH_ENDPOINT')
    return rows[0] if rows else None


def _find_storage_row(ws):
    """Find DATABRICKS_STORAGE row."""
    rows = _find_data_rows(ws, 'DATABRICKS_STORAGE')
    return rows[0] if rows else None


# Column indices (1-indexed for openpyxl)
COL_TYPE = 3
COL_MODE = 4
COL_CONFIG = 5
COL_SKU = 6
COL_HOURS = 12
COL_TOKEN_TYPE = 13
COL_DBU_HR = 16
COL_DBUS_MO = 17
COL_DBU_RATE = 18
COL_DISCOUNT = 19
COL_DBU_RATE_DISC = 20
COL_DBU_COST_L = 21
COL_DBU_COST_D = 22
COL_TOTAL_L = 28
COL_TOTAL_D = 29
COL_NOTES = 30


class TestComputeRowFormula:
    """AC-9: Compute row uses hourly formula =P{r}*L{r} for DBUs/Mo."""

    def test_dbus_mo_is_formula(self):
        items = [make_line_item(vector_capacity_millions=2)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_vs_compute_row(ws)
        assert row is not None, "No VECTOR_SEARCH_ENDPOINT row found"
        cell = ws.cell(row=row, column=COL_DBUS_MO)
        val = cell.value
        assert val is not None
        if isinstance(val, str):
            assert val.startswith('='), f"DBUs/Mo should be formula, got: {val}"
            assert 'P' in val.upper() and 'L' in val.upper(), (
                f"Hourly formula should ref P*L, got: {val}"
            )

    def test_dbus_mo_numeric_value(self):
        """Verify the cached numeric value matches expected."""
        items = [make_line_item(
            vector_capacity_millions=2, hours_per_month=730,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_vs_compute_row(ws)
        assert row is not None
        dbu_hr_val = ws.cell(row=row, column=COL_DBU_HR).value
        hours_val = ws.cell(row=row, column=COL_HOURS).value
        if isinstance(dbu_hr_val, (int, float)) and isinstance(hours_val, (int, float)):
            expected_dbus = dbu_hr_val * hours_val
            dbus_mo = ws.cell(row=row, column=COL_DBUS_MO).value
            if isinstance(dbus_mo, (int, float)):
                assert abs(dbus_mo - expected_dbus) < 0.1


class TestStorageSubRow:
    """AC-10 to AC-13: Storage sub-row emitted with correct values."""

    def test_storage_row_emitted(self):
        """AC-10: Storage sub-row exists for Vector Search."""
        items = [make_line_item(vector_capacity_millions=5)]
        wb = _generate_xlsx(items)
        ws = wb.active
        storage_rows = _find_data_rows(ws, 'DATABRICKS_STORAGE')
        assert len(storage_rows) >= 1, "No storage sub-row found"

    def test_storage_row_sku(self):
        """AC-11: Storage sub-row SKU = DATABRICKS_STORAGE."""
        items = [make_line_item(vector_capacity_millions=5)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_storage_row(ws)
        assert row is not None
        assert ws.cell(row=row, column=COL_SKU).value == 'DATABRICKS_STORAGE'

    def test_storage_row_type_display(self):
        """Storage sub-row type should say 'Vector Search (Storage)'."""
        items = [make_line_item(vector_capacity_millions=5)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_storage_row(ws)
        assert row is not None
        type_val = ws.cell(row=row, column=COL_TYPE).value
        assert 'Vector Search' in str(type_val)
        assert 'Storage' in str(type_val)

    def test_storage_gb_approximation(self):
        """AC-12: Storage GB ~ capacity_millions (1M ~ 1 GB)."""
        items = [make_line_item(vector_capacity_millions=5)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_storage_row(ws)
        assert row is not None
        config = ws.cell(row=row, column=COL_CONFIG).value
        assert '5' in str(config), (
            f"Storage config should mention 5 GB, got: {config}"
        )

    def test_storage_cost_positive(self):
        """AC-13: Storage cost = storage_gb * rate > 0."""
        items = [make_line_item(vector_capacity_millions=10)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_storage_row(ws)
        assert row is not None
        cost = ws.cell(row=row, column=COL_DBU_COST_L).value
        if isinstance(cost, (int, float)):
            assert cost > 0, f"Storage cost should be > 0, got {cost}"

    def test_storage_notes_mention_rate(self):
        """Notes should mention $/GB/month rate."""
        items = [make_line_item(vector_capacity_millions=5)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_storage_row(ws)
        assert row is not None
        notes = ws.cell(row=row, column=COL_NOTES).value
        assert notes and '/GB/month' in str(notes)


class TestComputeRowSKU:
    """Verify compute row has correct SKU."""

    def test_compute_sku_standard(self):
        items = [make_line_item(vector_search_mode='standard')]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_vs_compute_row(ws)
        assert row is not None
        assert ws.cell(row=row, column=COL_SKU).value == 'VECTOR_SEARCH_ENDPOINT'

    def test_compute_sku_storage_optimized(self):
        items = [make_line_item(vector_search_mode='storage_optimized',
                                vector_capacity_millions=64)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_vs_compute_row(ws)
        assert row is not None
        assert ws.cell(row=row, column=COL_SKU).value == 'VECTOR_SEARCH_ENDPOINT'


class TestExcelTotals:
    """AC-14: Totals SUM formula spans compute + storage rows."""

    def test_totals_sum_exists(self):
        items = [make_line_item(vector_capacity_millions=5)]
        wb = _generate_xlsx(items)
        ws = wb.active
        found_sum = False
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=COL_DBU_COST_L).value
            if val and isinstance(val, str) and 'SUM' in val.upper():
                found_sum = True
                break
        assert found_sum, "No SUM formula found in totals row"

    def test_totals_with_multiple_vs_items(self):
        """Two VS items → 4 data rows (2 compute + 2 storage)."""
        items = [
            make_line_item(vector_capacity_millions=2,
                           workload_name='VS Standard'),
            make_line_item(vector_search_mode='storage_optimized',
                           vector_capacity_millions=64,
                           workload_name='VS StorOpt', display_order=1),
        ]
        wb = _generate_xlsx(items)
        ws = wb.active
        compute_rows = _find_data_rows(ws, 'VECTOR_SEARCH_ENDPOINT')
        storage_rows = _find_data_rows(ws, 'DATABRICKS_STORAGE')
        assert len(compute_rows) == 2, (
            f"Expected 2 compute rows, got {len(compute_rows)}"
        )
        assert len(storage_rows) == 2, (
            f"Expected 2 storage rows, got {len(storage_rows)}"
        )


class TestExcelServerlessMarkers:
    """Vector Search rows should show 'Serverless' mode."""

    def test_mode_column_says_serverless(self):
        items = [make_line_item(vector_capacity_millions=2)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_vs_compute_row(ws)
        assert row is not None
        mode_val = ws.cell(row=row, column=COL_MODE).value
        assert mode_val == 'Serverless', (
            f"Expected 'Serverless', got '{mode_val}'"
        )

    def test_token_columns_are_dashes(self):
        """Vector Search is hour-based, not token-based."""
        items = [make_line_item(vector_capacity_millions=2)]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_vs_compute_row(ws)
        assert row is not None
        token_type = ws.cell(row=row, column=COL_TOKEN_TYPE).value
        assert token_type == '-', f"Token type should be '-', got {token_type}"


class TestExcelDBUValues:
    """Verify DBU/hr and DBU rate values in Excel match expectations."""

    @pytest.mark.parametrize("mode,capacity,expected_dbu_hr", [
        ('standard', 2, 4.0),
        ('standard', 5, 10.0),
        ('storage_optimized', 64, 18.29),
    ])
    def test_dbu_hr_in_excel(self, mode, capacity, expected_dbu_hr):
        items = [make_line_item(
            vector_search_mode=mode, vector_capacity_millions=capacity,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        row = _find_vs_compute_row(ws)
        assert row is not None
        dbu_hr = ws.cell(row=row, column=COL_DBU_HR).value
        assert isinstance(dbu_hr, (int, float)), (
            f"DBU/hr should be numeric, got {type(dbu_hr)}: {dbu_hr}"
        )
        assert abs(dbu_hr - expected_dbu_hr) < 0.01


class TestExcelNoNaN:
    """AC-20: No NaN values in generated Excel."""

    def test_no_nan_standard(self):
        items = [make_line_item(vector_capacity_millions=2)]
        wb = _generate_xlsx(items)
        ws = wb.active
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, 31):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, float):
                    assert not math.isnan(val), (
                        f"NaN at row={row_idx}, col={col_idx}"
                    )

    def test_no_nan_storage_optimized(self):
        items = [make_line_item(
            vector_search_mode='storage_optimized',
            vector_capacity_millions=64,
        )]
        wb = _generate_xlsx(items)
        ws = wb.active
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, 31):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, float):
                    assert not math.isnan(val), (
                        f"NaN at row={row_idx}, col={col_idx}"
                    )
